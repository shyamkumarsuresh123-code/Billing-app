import os
from datetime import date
from flask import Flask, render_template, request, send_file, jsonify
from openpyxl import load_workbook
from num2words import num2words

# --- Configuration ---
app = Flask(__name__)
DATABASE_FILE = "invoices_db.xlsx"
TEMPLATE_FILE = "1stop.xlsx"


# --- This is your original code, adapted for Flask ---
def generate_excel_bill(invoice_data, items_list):
    """
    Fills the Excel template with invoice data.
    Takes a dictionary of invoice data and a list of item dictionaries.
    """
    try:
        # Unpack the invoice data dictionary
        inv_num = invoice_data['inv_num']
        inv_date = invoice_data['inv_date']
        order_num = invoice_data['order_num']
        order_date = invoice_data['order_date']
        bill_type = invoice_data['bill_type']
        cust_name = invoice_data['cust_name']
        cust_addr = invoice_data['cust_addr']
        cust_phone = invoice_data['cust_phone']
        cust_gstin = invoice_data['cust_gstin']
        cust_state = invoice_data['cust_state']

        total_invoice_amount = invoice_data['total_invoice_amount']
        total_cgst_amount = invoice_data['total_cgst_amount']
        total_sgst_amount = invoice_data['total_sgst_amount']

        tpl_wb = load_workbook(TEMPLATE_FILE)
        tpl_ws = tpl_wb.active

        # --- Fill Customer/Invoice Details ---
        tpl_ws['A10'] = cust_name
        tpl_ws['A11'] = cust_addr
        tpl_ws['A12'] = cust_phone
        tpl_ws['H2'] = inv_num
        tpl_ws['K2'] = inv_date
        tpl_ws['J5'] = order_num
        tpl_ws['J6'] = order_date
        tpl_ws['J13'] = cust_gstin
        tpl_ws['J14'] = cust_state
        tpl_ws['I4'] = bill_type

        # --- Fill Total Amounts ---
        tpl_ws['L36'] = total_invoice_amount
        tpl_ws['C38'] = total_cgst_amount
        tpl_ws['C39'] = total_sgst_amount

        amount_in_words = num2words(total_invoice_amount, to='currency', lang='en_IN').title()
        tpl_ws['A44'] = f"Rupees {amount_in_words} Only"

        # --- Fill Item Table ---
        start_row = 20  # Your original start row

        for i, item in enumerate(items_list):
            row = start_row + i

            # Get data from the item dictionary
            desc = item['desc']
            hsn = item['hsn']
            qty = float(item['qty'])
            rate = float(item['rate'])
            gst_percent = float(item['gst'])
            item_total = float(item['total'])

            base_cost = qty * rate
            cgst_percent = gst_percent / 2
            sgst_percent = gst_percent / 2
            item_cgst_amount = base_cost * (cgst_percent / 100)
            item_sgst_amount = base_cost * (sgst_percent / 100)

            # --- Write to cells (based on your original code) ---
            tpl_ws[f'A{row}'] = i + 1
            tpl_ws[f'B{row}'] = desc
            tpl_ws[f'C{row}'] = hsn
            tpl_ws[f'D{row}'] = qty
            tpl_ws[f'E{row}'] = rate
            tpl_ws[f'G{row}'] = cgst_percent
            tpl_ws[f'H{row}'] = item_cgst_amount
            tpl_ws[f'I{row}'] = sgst_percent
            tpl_ws[f'J{row}'] = item_sgst_amount
            tpl_ws[f'K{row}'] = item_total

        output_filename = f"Invoice_{inv_num}.xlsx"
        tpl_wb.save(output_filename)
        return output_filename

    except Exception as e:
        print(f"Error generating Excel: {e}")
        raise


def save_to_database(invoice_data, items_list):
    """
    Saves the invoice data to the 'invoices_db.xlsx' file.
    """
    try:
        db_wb = load_workbook(DATABASE_FILE)
        db_ws = db_wb.active

        # Convert items list to the pipe-and-comma format
        # item = (desc, hsn, qty, rate, gst, total)
        items_str_list = []
        for item in items_list:
            item_tuple = (item['desc'], item['hsn'], item['qty'], item['rate'], item['gst'], item['total'])
            items_str_list.append(",".join(map(str, item_tuple)))
        items_str = "|".join(items_str_list)

        # Create the row to save
        data_to_save = [
            invoice_data['inv_num'],
            invoice_data['inv_date'],
            invoice_data['order_num'],
            invoice_data['order_date'],
            invoice_data['bill_type'],
            invoice_data['cust_name'],
            invoice_data['cust_addr'],
            invoice_data['cust_phone'],
            invoice_data['cust_gstin'],
            invoice_data['cust_state'],
            items_str,
            invoice_data['total_invoice_amount'],
            invoice_data['total_cgst_amount'],
            invoice_data['total_sgst_amount']
        ]

        db_ws.append(data_to_save)
        db_wb.save(DATABASE_FILE)

    except Exception as e:
        print(f"Error saving to database: {e}")
        raise


# --- Web Routes ---

@app.route('/')
def index():
    """Serves the main HTML page."""
    return render_template('index.html')


@app.route('/save_invoice', methods=['POST'])
def save_invoice():
    """
    Receives invoice data from the browser, saves it,
    generates the bill, and sends it back for download.
    """
    try:
        data = request.json
        customer_details = data['customerDetails']
        items_list = data['items']  # This is a list of item dictionaries

        # --- Calculate totals (like your original save_invoice) ---
        total_invoice_amount = 0.0
        total_cgst_amount = 0.0
        total_sgst_amount = 0.0

        for item in items_list:
            qty = float(item['qty'])
            rate = float(item['rate'])
            gst_percent = float(item['gst'])

            base_cost = qty * rate
            total_invoice_amount += float(item['total'])
            total_cgst_amount += base_cost * ((gst_percent / 2) / 100)
            total_sgst_amount += base_cost * ((gst_percent / 2) / 100)

        # Add calculated totals to the customer_details dictionary
        customer_details['total_invoice_amount'] = total_invoice_amount
        customer_details['total_cgst_amount'] = total_cgst_amount
        customer_details['total_sgst_amount'] = total_sgst_amount

        # 1. Save to Database
        save_to_database(customer_details, items_list)

        # 2. Generate Excel Bill
        output_file = generate_excel_bill(customer_details, items_list)

        # 3. Send the file to the user for download
        return send_file(output_file, as_attachment=True)

    except Exception as e:
        print(f"Error in /save_invoice: {e}")
        return jsonify({"error": str(e)}), 500


# --- Run the Application ---
if __name__ == "__main__":
    # Make sure your database and template files exist
    if not os.path.exists(DATABASE_FILE):
        print(f"Warning: '{DATABASE_FILE}' not found. It will be created on first save.")
    if not os.path.exists(TEMPLATE_FILE):
        print(f"CRITICAL ERROR: Template file '{TEMPLATE_FILE}' not found. Exiting.")
        exit()

    app.run(debug=True)